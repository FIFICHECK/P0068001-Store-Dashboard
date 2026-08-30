#!/usr/bin/env python3
"""
P0068001 Store Dashboard — data processor
1. Update order_reports_manifest.json (newest-first)
2. Generate monthly merged reports (reports/monthly/)
3. Generate sales_trend_data.js (by SKU by daily GMV & QTY)
"""
import json
import os
import re
import openpyxl
from datetime import datetime

BASE = os.environ.get("P0068001_BASE", "/home/snkwok/P0068001-Store-Dashboard")
# B pilot (2026-08-29): data/reports 移去 private repo (dashboard-private-data)
# cron 用 P0068001_BASE=~/dashboard-private-data/P0068001 寫入 private repo；default 行為不變
REPORTS_DIR = os.path.join(BASE, "reports", "order_reports")
MONTHLY_DIR = os.path.join(BASE, "reports", "monthly")
DATA_DIR = os.path.join(BASE, "data")
MANIFEST_PATH = os.path.join(DATA_DIR, "order_reports_manifest.json")
TREND_PATH = os.path.join(DATA_DIR, "sales_trend_data.js")

os.makedirs(MONTHLY_DIR, exist_ok=True)


def parse_filename(fname):
    """ECOM-MMSNG_DAILY_ORDER_P0068001_20260802235959.xlsx -> (date_iso, timestamp)"""
    m = re.search(r"_(\d{8})(\d{6})\.xlsx$", fname)
    if not m:
        return None, None
    ds = m.group(1)
    date_iso = f"{ds[:4]}-{ds[4:6]}-{ds[6:8]}"
    ts = f"{m.group(2)[:2]}:{m.group(2)[2:4]}"
    return date_iso, ts


def extract_stats(path):
    """Return (gmv, order_lines, qty) from a daily order report xlsx."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    gmv = 0.0
    lines = 0
    qty = 0
    for r in range(5, ws.max_row + 1):
        v = ws.cell(row=r, column=27).value
        if isinstance(v, (int, float)):
            gmv += v
            lines += 1
        q = ws.cell(row=r, column=24).value
        if isinstance(q, (int, float)):
            qty += q
    return gmv, lines, int(qty)


def read_orders(path):
    """Read per-SKU order lines: list of (date_iso, sku_id, brand_cn, sku_name, qty, gmv)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for r in range(5, ws.max_row + 1):
        sku = ws.cell(row=r, column=18).value
        if sku is None:
            continue
        sku_str = str(sku).strip()
        # Skip header/legend rows that leak into data (e.g. 'SKU ID')
        if sku_str.lower() in ("sku id", "sku", "sku_id") or sku_str.startswith("SKU"):
            continue
        qty = ws.cell(row=r, column=24).value or 0
        gmv = ws.cell(row=r, column=27).value or 0
        if not isinstance(qty, (int, float)):
            qty = 0
        if not isinstance(gmv, (int, float)):
            gmv = 0
        name_cn = ws.cell(row=r, column=22).value or ws.cell(row=r, column=21).value or sku_str
        brand_cn = ws.cell(row=r, column=20).value or ws.cell(row=r, column=19).value or ""
        # Normalize SKU: prepend store prefix P0068001_S_ if not already present
        if not sku_str.startswith("P0068001_S_"):
            sku_str = "P0068001_S_" + sku_str
        rows.append((sku_str, str(brand_cn), str(name_cn), float(qty), float(gmv)))
    return rows


def build_monthly_report(month_key, files):
    """Merge all daily reports of one month into a single xlsx."""
    if not files:
        return None
    out_path = os.path.join(MONTHLY_DIR, f"ECOM-MMSNG_DAILY_ORDER_P0068001_{month_key.replace('-', '')}_MONTHLY.xlsx")
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "hktv_order_list"

    first_wb = openpyxl.load_workbook(files[0], data_only=True)
    first_ws = first_wb.active

    # Copy header rows 1-4 from first file
    for r in range(1, 5):
        for c in range(1, first_ws.max_column + 1):
            ws_out.cell(row=r, column=c).value = first_ws.cell(row=r, column=c).value
    # Fix print date header
    ws_out.cell(row=2, column=2).value = datetime.now().strftime("%Y-%b-%d %H:%M:%S")

    # Copy data rows (5+) from all files — skip the header row (row 5 in each
    # daily file is "Delivery Mode | Warehouse ID | ...") and any non-data rows.
    # Same rule as extract_stats(): only rows with numeric GMV in col AA (27)
    # count as order lines. (Fixed 2026-08-28: monthly merged previously
    # included 1 duplicated header per daily file → order_lines inflated.)
    out_row = 5
    total_gmv = 0.0
    total_qty = 0.0
    order_lines = 0
    for fpath in sorted(files):
        wb = openpyxl.load_workbook(fpath, data_only=True)
        ws = wb.active
        for r in range(5, ws.max_row + 1):
            v27 = ws.cell(row=r, column=27).value
            if not isinstance(v27, (int, float)):
                continue
            vals = []
            for c in range(1, ws.max_column + 1):
                vals.append(ws.cell(row=r, column=c).value)
            for c, v in enumerate(vals, start=1):
                ws_out.cell(row=out_row, column=c).value = v
            g = v27
            q = vals[23] if len(vals) > 23 and isinstance(vals[23], (int, float)) else 0.0
            total_gmv += g
            total_qty += q
            order_lines += 1
            out_row += 1

    # Update summary cells (row 2)
    ws_out.cell(row=2, column=6).value = round(total_gmv, 2)
    ws_out.cell(row=1, column=1).value = "Daily Order Report (Monthly Merged)"

    wb_out.save(out_path)
    return out_path, round(total_gmv, 2), order_lines, total_qty


def main():
    # Collect all daily reports
    files = sorted(
        os.path.join(REPORTS_DIR, f)
        for f in os.listdir(REPORTS_DIR)
        if f.startswith("ECOM-MMSNG_DAILY_ORDER_P0068001_") and f.endswith(".xlsx")
    )
    print(f"Found {len(files)} daily reports")

    # Group by date, prefer 23:59 (final) per date
    by_date = {}
    for fpath in files:
        fname = os.path.basename(fpath)
        date_iso, ts = parse_filename(fname)
        if not date_iso:
            continue
        if date_iso not in by_date or ts == "23:59":
            by_date[date_iso] = fpath

    # Build manifest (newest first)
    manifest = []
    for date_iso, fpath in sorted(by_date.items(), reverse=True):
        fname = os.path.basename(fpath)
        _, ts = parse_filename(fname)
        gmv, lines, qty = extract_stats(fpath)
        month_day = f"{int(date_iso[5:7])}月{int(date_iso[8:10])}日"
        manifest.append({
            "date": month_day,
            "date_iso": date_iso,
            "gmv": f"${gmv:,.2f}",
            "gmv_raw": round(gmv, 2),
            "orders": lines,
            "qty": qty,
            "timestamp": ts,
            "filename": fname,
        })
    with open(MANIFEST_PATH, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)
    print(f"Manifest: {len(manifest)} entries, latest {manifest[0]['date_iso']} GMV {manifest[0]['gmv']}")

    # Build monthly merged reports
    by_month = {}
    for date_iso, fpath in by_date.items():
        mk = date_iso[:7]
        by_month.setdefault(mk, []).append(fpath)

    monthly_meta = []
    for mk in sorted(by_month.keys()):
        res = build_monthly_report(mk, by_month[mk])
        if res:
            out_path, gmv, lines, qty = res
            fname = os.path.basename(out_path)
            monthly_meta.append({
                "month": mk,
                "label": f"{int(mk[5:7])}月 {mk[:4]}",
                "gmv": f"${gmv:,.2f}",
                "orders": lines,
                "qty": qty,
                "days": len(by_month[mk]),
                "filename": fname,
            })
            print(f"Monthly {mk}: {len(by_month[mk])} days, GMV ${gmv:,.2f}, {lines} lines, {qty} qty")

    monthly_meta.sort(key=lambda x: x["month"], reverse=True)

    # Preserve monthly-only entries (imported via import_monthly_csv.py, e.g. 2026-01..07)
    # These have no daily XLSX so they are not in by_month; merge them back in.
    existing_path = os.path.join(DATA_DIR, "monthly_reports.json")
    if os.path.exists(existing_path):
        try:
            with open(existing_path, encoding="utf-8") as f:
                existing_meta = json.load(f)
            existing_months = {m["month"] for m in monthly_meta}
            for em in existing_meta:
                if em["month"] not in existing_months and not em.get("filename"):
                    monthly_meta.append(em)
            monthly_meta.sort(key=lambda x: x["month"], reverse=True)
        except Exception as e:
            print(f"Warning: could not merge existing monthly_reports.json: {e}")

    with open(existing_path, "w", encoding="utf-8") as f:
        json.dump(monthly_meta, f, ensure_ascii=False, indent=2)
    print(f"Monthly reports: {len(monthly_meta)} entries")

    # Build Sales Trend data (by SKU by daily GMV & QTY)
    sku_data = {}   # sku_id -> {brand, name, by_date: {date: [gmv, qty]}}
    dates = sorted(by_date.keys())
    for date_iso, fpath in by_date.items():
        for sku, brand, name, qty, gmv in read_orders(fpath):
            if sku not in sku_data:
                sku_data[sku] = {"brand": brand, "name": name, "by_date": {}}
            d = sku_data[sku]["by_date"]
            if date_iso not in d:
                d[date_iso] = [0.0, 0]
            d[date_iso][0] += gmv
            d[date_iso][1] += qty

    # Build chart arrays
    sku_list = []
    for sku, info in sku_data.items():
        gmv_arr = []
        qty_arr = []
        for dt in dates:
            v = info["by_date"].get(dt, [0, 0])
            gmv_arr.append(round(v[0], 2))
            qty_arr.append(v[1])
        sku_list.append({
            "sku": sku,
            "brand": info["brand"],
            "name": info["name"],
            "gmv": gmv_arr,
            "qty": qty_arr,
        })
    # Sort by total GMV desc
    sku_list.sort(key=lambda s: sum(s["gmv"]), reverse=True)

    trend = {
        "dates": dates,
        "skus": sku_list,
    }
    with open(TREND_PATH, "w", encoding="utf-8") as f:
        f.write("window.salesTrendData = ")
        json.dump(trend, f, ensure_ascii=False)
        f.write(";\n")
    # B pilot fix (2026-08-30): 前端 TREND_URL 讀 sales_trend_data.json — 一併生成
    TREND_JSON_PATH = os.path.join(DATA_DIR, "sales_trend_data.json")
    with open(TREND_JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(trend, f, ensure_ascii=False)
    print(f"Sales Trend: {len(sku_list)} SKUs, {len(dates)} dates")


if __name__ == "__main__":
    main()
