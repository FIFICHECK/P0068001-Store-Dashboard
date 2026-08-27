#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""P0068001 Shelf Life Check — API mode (recommended) or xlsx mode.

API mode (cron-friendly, no async UI export needed):
  1. Login MMS -> accessToken (passed via --token or --token-b64)
  2. POST https://merchant-web.shoalter.com/product/v2/products
     body {"page":N,"size":200,"buCode":["HKTV"],"disableInventory":true,
           "orderBy":["modified_time desc"],"storeId":[<numeric store id>]}
     page size cap = 200; 22,826 SKUs -> 115 pages
  3. Per product: additional.hktv.store_sku_id (P0068001_S_xxx), minimum_shelf_life,
     additional.hktv.primary_category_code
  4. Compare vs GreenLab.xlsx Sheet4 Column I requirement:
     - primary_category_code must EXIST in GreenLab mapping FIRST (user rule 1)
     - SKU DP < GreenLab required            -> WARN (below)
     - SKU DP empty (null) + GreenLab has req -> WARN (missing)
     - SKU DP >= GreenLab required           -> OK (not listed)
     - GreenLab no req (— / Exclude in GL / unmatched) -> not checked

Output: data/shelf_life_check.json + printed summary.

Usage:
  python check_shelf_life.py --api --token-b64 <b64> [--store-id 33249]
  python check_shelf_life.py --api --token <raw_jwt>
  python check_shelf_life.py --exports F1.xlsx [F2.xlsx ...]   # xlsx mode (fallback)
  Options: --greenlab PATH --out PATH
"""
import argparse
import base64
import datetime
import json
import os
import re
import subprocess
import sys
import warnings

import openpyxl

warnings.filterwarnings('ignore')

# P0068001
STORE_ID = 33249          # numeric store id (from storefront dropdown checkbox value)
STORE_CODE = 'P0068001'
PRODUCT_API = 'https://merchant-web.shoalter.com/product/v2/products'
PAGE_SIZE = 200

COL_SKU = 1
COL_CAT = 3
COL_BRAND = 4
COL_NAME_CHI = 12
COL_DP = 119

NUM_RE = re.compile(r'(\d+)')


def parse_dp(val):
    """Extract days from Minimum Shelf Life value (int / '46' / '至少91日食用期')."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ('', '—', '-', 'None', 'N/A', 'NA'):
        return None
    m = NUM_RE.search(s)
    return int(m.group(1)) if m else None


def parse_greenlab_period(val):
    if val is None:
        return None
    s = str(val).strip()
    if s in ('', '—', '-', 'None', 'N/A', 'Exclude in GL', 'exclude'):
        return None
    m = NUM_RE.search(s)
    return int(m.group(1)) if m else None


def load_greenlab(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb['Sheet4']
    it = ws.iter_rows(values_only=True)
    next(it)
    gl = {}
    for r in it:
        if r[4] is None:
            continue
        gl[str(r[4]).strip()] = (r[8], r[5], r[6])
    wb.close()
    return gl


def api_login_token(token_b64=None, token=None):
    if token_b64:
        return base64.b64decode(token_b64).decode()
    return token


def api_fetch_products(token, store_id):
    """Loop all pages of /product/v2/products for P0068001. Returns list of dicts."""
    headers = [
        '-H', 'Content-Type: application/json',
        '-H', f'Authorization: Bearer {token}',
        '-H', 'Accept: application/json',
        '-H', 'Origin: https://merchant.shoalter.com',
        '-H', 'Referer: https://merchant.shoalter.com/',
    ]
    all_products = []
    page = 1
    total = None
    while True:
        body = json.dumps({
            'page': page, 'size': PAGE_SIZE, 'buCode': ['HKTV'],
            'disableInventory': True, 'orderBy': ['modified_time desc'],
            'storeId': [store_id],
        })
        cmd = ['curl', '-s', '-X', 'POST', PRODUCT_API] + headers + ['-d', body]
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
        try:
            j = json.loads(r.stdout)
        except Exception:
            print(f'  page {page}: JSON parse fail, raw={r.stdout[:150]}')
            break
        if j.get('status') != 1 or not j.get('data'):
            print(f'  page {page}: FAIL {json.dumps(j, ensure_ascii=False)[:200]}')
            break
        d = j['data']
        if total is None:
            total = d.get('totalElements')
            print(f'  total products: {total} ({d.get("totalPages")} pages @ {PAGE_SIZE})')
        prods = d.get('products', [])
        all_products.extend(prods)
        if not d.get('last') and page < 200:
            page += 1
        else:
            break
        if page % 20 == 0:
            print(f'  ... fetched {len(all_products)} products (page {page})')
    print(f'  fetched {len(all_products)} products')
    return all_products


def match_suggested(cat, suggested_map):
    """Longest-prefix match of primary_category_code against MUJI Suggested AA-group map.
    Returns (days, matched_prefix) or (None, None)."""
    if not cat or not suggested_map:
        return None, None
    # try longest prefix first (AA1115300 before AA1115)
    for length in (12, 11, 10, 9, 8, 7, 6, 5, 4):
        if len(cat) < length:
            continue
        pfx = cat[:length]
        if pfx in suggested_map:
            return suggested_map[pfx], pfx
    return None, None


def check_rows(rows, gl, suggested_map=None):
    """rows: list of (sku_display, name, online_status, cat, dp). Returns (warnings_list, total, checked)."""
    total = 0
    checked = 0
    warnings_list = []
    seen = set()
    for sku_display, name, online_status, cat_raw, dp in rows:
        if not sku_display or sku_display in seen:
            continue
        seen.add(sku_display)
        total += 1
        cat = (cat_raw or '').split(',')[0].strip()
        # Rule 1: primary_category_code must match GreenLab primary_category_code FIRST
        gl_entry = gl.get(cat) if cat else None
        if gl_entry is None:
            continue  # unmatched category -> not checked
        gl_period = parse_greenlab_period(gl_entry[0])
        if gl_period is None:
            continue  # GreenLab no requirement
        checked += 1
        # Rule 2: 建議到期日 = MUJI Suggested AA-group value (longest prefix), fallback GreenLab + 1
        sug_days, sug_pfx = match_suggested(cat, suggested_map)
        if sug_days is None:
            sug_days = gl_period + 1
            sug_pfx = ''
        if dp is None:
            warnings_list.append({
                'sku': sku_display,
                'name': name or '',
                'online_status': online_status or '',
                'category_code': cat,
                'dp': '',
                'greenlab_required': gl_period,
                'suggested': sug_days,
                'suggested_label': f'至少{sug_days}日食用期' + (f'（{sug_pfx} 建議）' if sug_pfx else ''),
                'status': 'missing',
            })
        elif dp < gl_period:
            warnings_list.append({
                'sku': sku_display,
                'name': name or '',
                'online_status': online_status or '',
                'category_code': cat,
                'dp': dp,
                'greenlab_required': gl_period,
                'suggested': sug_days,
                'suggested_label': f'至少{sug_days}日食用期' + (f'（{sug_pfx} 建議）' if sug_pfx else ''),
                'status': 'below',
            })
    return warnings_list, total, checked


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--api', action='store_true', help='API mode (recommended)')
    ap.add_argument('--token', default=None, help='raw JWT accessToken')
    ap.add_argument('--token-b64', default=None, help='base64-encoded JWT accessToken')
    ap.add_argument('--store-id', type=int, default=STORE_ID)
    ap.add_argument('--exports', nargs='+', default=None)
    ap.add_argument('--greenlab', default='/home/snkwok/P0068001-Store-Dashboard/data/GreenLab.xlsx')
    ap.add_argument('--out', default='/home/snkwok/P0068001-Store-Dashboard/data/shelf_life_check.json')
    a = ap.parse_args()
    store_id = a.store_id

    gl = load_greenlab(a.greenlab)
    print(f'GreenLab mapping: {len(gl)} categories')

    # MUJI Suggested AA-group -> days map (data/suggested_shelf_life.json)
    suggested_map = {}
    sug_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'data', 'suggested_shelf_life.json')
    try:
        with open(sug_path, encoding='utf-8') as _f:
            sug_raw = json.load(_f)
        suggested_map = {k: v for k, v in sug_raw.items() if not k.startswith('_') and isinstance(v, int)}
        print(f'MUJI Suggested map: {len(suggested_map)} AA groups')
    except Exception as e:
        print(f'  (no suggested map: {e})')

    warnings_list, total, checked = [], 0, 0
    source = ''

    if a.api:
        token = api_login_token(a.token_b64, a.token)
        if not token:
            print('ERROR: --api requires --token or --token-b64')
            sys.exit(1)
        products = api_fetch_products(token, store_id)
        rows = []
        for p in products:
            hktv = (p.get('additional') or {}).get('hktv') or {}
            rows.append((
                hktv.get('store_sku_id') or p.get('sku_id'),
                p.get('sku_name_ch') or p.get('sku_name_en') or '',
                hktv.get('online_status') or '',
                hktv.get('primary_category_code') or '',
                parse_dp(p.get('minimum_shelf_life')),
            ))
        source = f'MMS API {datetime.date.today().isoformat()}'
        warnings_list, total, checked = check_rows(rows, gl, suggested_map)
    else:
        export_files = a.exports
        if not export_files:
            for pat in ['reports/product_exports/Product_export_all_column_*.xlsx',
                        '~/Downloads/Product_export_all_column_*.xlsx']:
                hits = sorted(__import__('glob').glob(os.path.expanduser(pat)))
                if hits:
                    export_files = hits
                    break
        if not export_files:
            print('ERROR: no export files (use --exports or --api)')
            sys.exit(1)
        rows = []
        for f in export_files:
            wb = openpyxl.load_workbook(f, read_only=True)
            ws = wb['Product Template']
            it = ws.iter_rows(values_only=True)
            hdr = next(it)
            if len(hdr) <= COL_DP or hdr[COL_DP] != 'Minimum Shelf Life':
                print(f'  SKIP {f}: DP col mismatch')
                wb.close()
                continue
            n = 0
            for r in it:
                if r[COL_SKU] is None:
                    continue
                rows.append((
                    f'{STORE_CODE}_S_{str(r[COL_SKU]).strip()}',
                    str(r[COL_NAME_CHI]).strip() if len(r) > COL_NAME_CHI and r[COL_NAME_CHI] else '',
                    str(r[COL_BRAND]).strip() if len(r) > COL_BRAND and r[COL_BRAND] else '',
                    str(r[COL_CAT]).strip() if r[COL_CAT] is not None else '',
                    parse_dp(r[COL_DP] if len(r) > COL_DP else None),
                ))
                n += 1
            print(f'  {os.path.basename(f)}: {n} rows')
            wb.close()
        source = os.path.basename(export_files[0]) if export_files else ''
        warnings_list, total, checked = check_rows(rows, gl, suggested_map)

    warnings_list.sort(key=lambda w: (w['status'], w['greenlab_required'], w['sku']))
    payload = {
        'generated_at': datetime.datetime.now().strftime('%Y-%m-%d %H:%M'),
        'source_export': source,
        'total_skus': total,
        'checked_skus': checked,
        'warning_count': len(warnings_list),
        'warnings': warnings_list,
    }
    os.makedirs(os.path.dirname(a.out), exist_ok=True)
    with open(a.out, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=1)
    print(f'Total SKUs: {total} | checked: {checked} | warnings: {len(warnings_list)}')
    print(f'  missing: {sum(1 for w in warnings_list if w["status"] == "missing")} | below: {sum(1 for w in warnings_list if w["status"] == "below")}')
    print('SAVED:', a.out)


if __name__ == '__main__':
    main()
